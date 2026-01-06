import openpyxl

def add_language_column(file_path):
    """
    Adds a 'language' column with default value 'eng_Latn' to the users.xlsx file.
    """
    try:
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Check if the 'language' column already exists
        headers = [cell.value for cell in sheet[1]]
        if 'language' not in headers:
            # Add the 'language' column header
            sheet.cell(row=1, column=len(headers) + 1, value='language')

            # Add 'eng_Latn' as the default value for all rows
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=len(headers) + 1, value='eng_Latn')

            # Save the workbook
            workbook.save(file_path)
            print("✅ 'language' column added successfully with default value 'eng_Latn'.")
        else:
            print("⚠️ 'language' column already exists.")

    except Exception as e:
        print(f"Error: {e}")


# Example usage
add_language_column("users.xlsx")
